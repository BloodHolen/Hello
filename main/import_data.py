import os
import sys
import django

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from django.contrib.auth.hashers import make_password
from main.models import User, Category, Manufacturer, Supplier, Unit, Product, PickupPoint, Order, OrderItem

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMPORT_DIR = r'C:\Users\Apocrypha\Desktop\Демоэкз\Demo_ex\Модуль 1\import'


def import_users():
    wb = openpyxl.load_workbook(os.path.join(IMPORT_DIR, 'user_import.xlsx'))
    ws = wb.active
    role_map = {
        'Администратор': 'admin',
        'Менеджер': 'manager',
        'Авторизированный клиент': 'client',
    }
    for row in list(ws.iter_rows(values_only=True))[1:]:
        if row[0] is None:
            continue
        role_ru, fio, login, password = row[0], row[1], row[2], str(row[3])
        fio_parts = fio.split()
        last_name = fio_parts[0] if len(fio_parts) > 0 else ''
        first_name = fio_parts[1] if len(fio_parts) > 1 else ''
        patronymic = fio_parts[2] if len(fio_parts) > 2 else ''
        role = role_map.get(role_ru, 'client')

        if not User.objects.filter(username=login).exists():
            User.objects.create(
                username=login,
                email=login,
                password=make_password(password),
                first_name=first_name,
                last_name=last_name,
                patronymic=patronymic,
                role=role,
            )
            print(f'User created: {login} ({role_ru})')
    print('Users import done.')


def import_products():
    wb = openpyxl.load_workbook(os.path.join(IMPORT_DIR, 'Tovar.xlsx'))
    ws = wb.active

    categories = set()
    manufacturers = set()
    suppliers = set()
    units = set()
    products_data = []

    for row in list(ws.iter_rows(values_only=True))[1:]:
        if row[0] is None:
            continue
        article, name, unit, price, supplier, manufacturer, category, discount, qty, desc, photo = row
        categories.add(category)
        manufacturers.add(manufacturer)
        suppliers.add(supplier)
        units.add(unit)
        products_data.append(row)

    for c in categories:
        Category.objects.get_or_create(name=c)
    for m in manufacturers:
        Manufacturer.objects.get_or_create(name=m)
    for s in suppliers:
        Supplier.objects.get_or_create(name=s)
    for u in units:
        Unit.objects.get_or_create(name=u)

    for row in products_data:
        article, name, unit, price, supplier, manufacturer, category, discount, qty, desc, photo = row
        unit_obj = Unit.objects.get(name=unit)
        supplier_obj = Supplier.objects.get(name=supplier)
        manufacturer_obj = Manufacturer.objects.get(name=manufacturer)
        category_obj = Category.objects.get(name=category)
        qty = int(qty) if qty else 0
        discount = int(discount) if discount else 0

        Product.objects.create(
            article=article,
            name=name,
            unit=unit_obj,
            price=price,
            supplier=supplier_obj,
            manufacturer=manufacturer_obj,
            category=category_obj,
            discount=discount,
            quantity=qty,
            description=desc or '',
            photo=f'products/{photo}' if photo else None,
        )
        print(f'Product created: {article}')
    print('Products import done.')


def import_pickup_points():
    wb = openpyxl.load_workbook(os.path.join(IMPORT_DIR, 'Пункты выдачи_import.xlsx'))
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        PickupPoint.objects.get_or_create(address=row[0])
    print('Pickup points import done.')


def import_orders():
    import datetime

    wb = openpyxl.load_workbook(os.path.join(IMPORT_DIR, 'Заказ_import.xlsx'))
    ws = wb.active

    pickup_points = list(PickupPoint.objects.all())

    all_users = list(User.objects.all())
    user_by_fio = {}
    for u in all_users:
        fio = f'{u.last_name} {u.first_name} {u.patronymic}'.strip()
        user_by_fio[fio] = u

    def parse_date(val):
        if val is None:
            return None
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        if isinstance(val, str):
            for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d.%m.%y']:
                try:
                    return datetime.datetime.strptime(val.strip(), fmt).date()
                except ValueError:
                    continue
        return None

    for row in list(ws.iter_rows(values_only=True))[1:]:
        if row[0] is None:
            continue
        order_num, composition, order_date, delivery_date, point_code, client_fio, order_code, status = row[:8]

        order_code = str(int(order_code))

        point_idx = point_code - 1 if isinstance(point_code, int) else 0
        if isinstance(point_code, str) and point_code.isdigit():
            point_idx = int(point_code) - 1
        pickup = pickup_points[point_idx] if point_idx < len(pickup_points) else None

        client = user_by_fio.get(client_fio.strip() if client_fio else '') if client_fio else None
        if not client:
            client = User.objects.filter(role='client').first()

        odate = parse_date(order_date)
        ddate = parse_date(delivery_date)

        if odate is None:
            print(f'Skipping order {order_code}: invalid order_date {order_date}')
            continue
        if ddate is None:
            print(f'Skipping order {order_code}: invalid delivery_date {delivery_date}')
            continue

        status_clean = str(status).strip() if status else 'новый'

        if not Order.objects.filter(code=order_code).exists():
            order = Order.objects.create(
                code=order_code,
                order_number=order_num,
                order_date=odate,
                delivery_date=ddate,
                pickup_point=pickup,
                client=client,
                status=status_clean,
            )
            print(f'Order created: {order_code}')

            parts = [p.strip() for p in composition.split(',')]
            for i in range(0, len(parts), 2):
                article = parts[i]
                qty = int(parts[i + 1]) if i + 1 < len(parts) else 1
                try:
                    product = Product.objects.get(article=article)
                    OrderItem.objects.create(order=order, product=product, quantity=qty)
                except Product.DoesNotExist:
                    print(f'Product {article} not found, skipping')
    print('Orders import done.')


if __name__ == '__main__':
    import_users()
    import_pickup_points()
    import_products()
    import_orders()
    print('All data imported successfully!')
